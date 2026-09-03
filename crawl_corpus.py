# -*- coding: utf-8 -*-
"""
按《02_搜索与试采样语料表》采集抖音视频详情与评论数据。

流程：
1. 读取 Excel「每日搜索记录」sheet，提取 20 条抖音分享短链接。
2. 短链接 → 重定向 → www.douyin.com/video/{aweme_id}。
3. DouyinAPI.get_work_info 抓视频详情（标题/发布时间/账号属性/统计）。
4. DouyinAPI.get_work_out_comment 抓每个视频前 N 条热评。
5. 用关键词规则对评论做分类标注（叙事类型/叙述者/异象类型等）。
6. 生成结果 Excel（视频记录 + 评论摘录两个 sheet）。

合规：不落盘用户名/头像/主页等可识别信息，仅保留账号类型与评论正文。
"""
import os
import re
import sys
import time
import warnings

warnings.filterwarnings('ignore')

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
os.chdir(BASE)

from datetime import datetime, timezone, timedelta

import requests
requests.packages.urllib3.disable_warnings()

import openpyxl

from utils.common_util import load_env
from dy_apis.douyin_api import DouyinAPI

# ── 路径配置 ──────────────────────────────────────────────────────────────
EXCEL_IN = os.path.join(os.path.dirname(BASE), '02_搜索与试采样语料表_v1(1).xlsx')
EXCEL_OUT = os.path.join(os.path.dirname(BASE), '02_搜索与试采样语料表_v1_采集结果.xlsx')

MAX_COMMENTS_PER_VIDEO = 100   # 每个视频抓取的热评数量
REQUEST_INTERVAL = 1.0         # 每次请求间隔（秒），降低风控
CN_TZ = timezone(timedelta(hours=8))


# ── 短链接解析 ────────────────────────────────────────────────────────────
def resolve_short_url(short_url: str):
    """短链接重定向到正式作品页，返回 (work_url, aweme_id)。"""
    r = requests.get(short_url, allow_redirects=True, timeout=20, verify=False,
                     headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'})
    final = r.url
    # 兼容视频 /video/{id} 与图文 /note/{id}
    m = re.search(r'/(?:video|note)/(\d+)', final)
    if m:
        aweme_id = m.group(1)
        return f'https://www.douyin.com/video/{aweme_id}', aweme_id
    m2 = re.search(r'modal_id=(\d+)', final)
    if m2:
        aweme_id = m2.group(1)
        return f'https://www.douyin.com/video/{aweme_id}', aweme_id
    raise RuntimeError(f'无法从短链接解析 aweme_id: {short_url} -> {final}')


def ts_to_date(ts) -> str:
    """Unix 秒级时间戳 → 北京时间日期字符串。"""
    try:
        ts = int(ts)
        if ts <= 0:
            return ''
        return datetime.fromtimestamp(ts, tz=CN_TZ).strftime('%Y-%m-%d')
    except Exception:
        return ''


def norm_date(v) -> str:
    """把表格里各种日期形态统一成 YYYY-MM-DD 字符串。"""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if not s:
        return ''
    # 2026/8.25 这种
    m = re.match(r'(\d{4})[/.](\d{1,2})(?:[.](\d{1,2}))?', s)
    if m:
        y = m.group(1)
        mo = m.group(2).zfill(2)
        d = (m.group(3) or '01').zfill(2)
        return f'{y}-{mo}-{d}'
    m2 = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m2:
        return f'{m2.group(1)}-{m2.group(2).zfill(2)}-{m2.group(3).zfill(2)}'
    return s


# ── 视频详情抓取（带重试，应对偶发风控）───────────────────────────────────
def fetch_work_info(auth, work_url: str, retries: int = 3):
    """抓取 aweme_detail，失败时刷新 msToken 并重试。"""
    last_err = ''
    for attempt in range(retries):
        try:
            resp = DouyinAPI.get_work_info(auth, work_url)
            detail = resp.get('aweme_detail') or {}
            if detail:
                return detail
            last_err = f'aweme_detail 为空 (resp keys: {list(resp.keys())[:5]})'
        except Exception as e:
            last_err = str(e)
        try:
            auth.refresh_mstoken()
        except Exception:
            pass
        time.sleep(2 * (attempt + 1))
    print(f'    [详情重试{retries}次仍失败] {last_err}')
    return None


# ── 评论抓取（限数量）─────────────────────────────────────────────────────
def get_comments_limited(auth, url: str, max_count: int):
    """抓取作品前 max_count 条一级评论（热评优先）。"""
    cursor = '0'
    comments = []
    last_cursor = None
    while len(comments) < max_count:
        try:
            resp = DouyinAPI.get_work_out_comment(auth, url, cursor)
        except Exception as e:
            print(f'    [评论抓取异常] cursor={cursor}: {e}')
            break
        batch = resp.get('comments') or []
        if not batch:
            break
        comments.extend(batch)
        cursor = str(resp.get('cursor', '0'))
        if resp.get('has_more', 0) != 1:
            break
        if cursor == last_cursor or cursor == '0':
            break
        last_cursor = cursor
        time.sleep(REQUEST_INTERVAL)
    return comments[:max_count]


# ── 账号类型分类 ──────────────────────────────────────────────────────────
def classify_account(author: dict) -> str:
    """粗略判断账号类型：个人 / 媒体 / 知识号 / 故事号 / 未知。"""
    if not author:
        return '未知'
    enterprise = (author.get('enterprise_verify_reason') or '').strip()
    custom = (author.get('custom_verify') or '').strip()
    signature = (author.get('signature') or '').strip()
    verify_type = author.get('verification_type', 0)

    if enterprise:
        return '媒体'
    if custom:
        # 领域创作者/医生/律师等垂直认证 → 知识号
        return '知识号'
    if any(k in signature for k in ['故事', '讲述', '说书', '评书', '传奇', '讲故']):
        return '故事号'
    if verify_type == 0:
        return '个人'
    return '个人'


# ── 关键词规则分类 ────────────────────────────────────────────────────────
# 叙事类型
_KW_NARRATIVE = {
    '预兆': ['托梦', '梦见', '梦到', '预兆', '征兆', '前兆', '托梦', '显灵', '暗示'],
    '报应': ['报应', '现世报', '因果', '遭报应', '恶有恶报', '善有善报', '积德', '缺德', '业障'],
    '灵异': ['鬼', '灵异', '闹鬼', '鬼魂', '怪事', '诡异', '阴森', '灵魂', '附体'],
    '禁忌劝诫': ['别不信', '不能', '不要', '禁忌', '忌讳', '千万别', '千万不能', '注意了'],
    '玩梗边界': ['哈哈哈', '笑死', '捂脸', '哈哈哈哈', '离谱', '绷不住'],
}

# 叙述者位置
_KW_NARRATOR = {
    '家人': ['我爸', '我妈', '我爷爷', '我奶奶', '我外公', '我外婆', '我姥姥', '我姥爷', '家里老人',
             '我叔', '我舅', '我姑', '我姨', '我哥', '我姐', '我妈', '我爸', '长辈'],
    '朋友': ['朋友', '同事', '同学', '邻居', '发小', '室友'],
    '地方传闻': ['我们村', '我们那', '我们这儿', '老家', '村里', '镇上', '县里', '我们当地'],
    '亲历': ['我', '本人', '自己', '亲身', '我家', '我亲眼', '我遇到', '我经历'],
    '泛称': ['听说', '据说', '老人说', '老一辈', '别人', '有人', '传说', '民间'],
}

# 证据标记
_KW_EVIDENCE = ['真的', '亲眼', '老人说', '老一辈说', '别不信', '真事', '亲身经历',
                '不骗你', '骗你是狗', '千真万确', '我发誓', '有图有真相']

# 异象/异常类型
_KW_OMEN = {
    '梦': ['梦', '托梦', '梦见'],
    '动物': ['狗', '猫', '鸟', '蛇', '龟', '鸡', '牛', '鱼', '乌鸦', '猫头鹰'],
    '声音': ['响声', '声音', '听到', '响动', '哭声', '笑声'],
    '器物': ['钟', '摆件', '照片', '镜子', '香', '蜡烛', '碗', '筷子'],
    '身体感应': ['鸡皮疙瘩', '心慌', '发冷', '头晕', '发麻', '发怵', '毛骨悚然', '背脊发凉'],
}

# 应验判断
_KW_FULFILL_YES = ['应验', '真的发生', '果然', '灵验', '实现', '成真', '兑现', '准了']
_KW_FULFILL_NO = ['没有', '不准', '没发生', '假的']

# 道德判断
_KW_MORAL = ['报应', '活该', '善恶', '积德', '缺德', '因果', '报', '天理', '老天有眼']

# 互动类型
_KW_INTERACT_QUERY = ['假的', '不信', '骗人', '吹牛', '迷信', '扯淡', '胡扯']
_KW_INTERACT_MEME = ['哈哈哈', '笑死', '捂脸', '哈哈哈哈', '离谱']
_KW_INTERACT_SUPPLEMENT = ['我也', '我也有', '我也是', '我也遇到', '+1', '我也是这样', '同款']


def _match_any(text: str, kws) -> bool:
    return any(k in text for k in kws)


def classify_narrative(text: str) -> str:
    hits = [k for k, kws in _KW_NARRATIVE.items() if _match_any(text, kws)]
    if not hits:
        return '灵异' if _match_any(text, ['怪', '事', '异']) else '其他'
    # 玩梗边界优先弱化
    if len(hits) >= 2:
        return '混合'
    return hits[0]


def classify_narrator(text: str) -> str:
    # 家人/朋友/地方传闻优先于泛称/亲历
    for k in ['家人', '朋友', '地方传闻']:
        if _match_any(text, _KW_NARRATOR[k]):
            return k
    if _match_any(text, _KW_NARRATOR['亲历']):
        return '亲历'
    if _match_any(text, _KW_NARRATOR['泛称']):
        return '泛称'
    return '泛称'


def classify_omen(text: str) -> str:
    hits = [k for k, kws in _KW_OMEN.items() if _match_any(text, kws)]
    if not hits:
        return ''
    return '/'.join(hits)


def classify_fulfillment(text: str) -> str:
    if _match_any(text, _KW_FULFILL_YES):
        return '是'
    if _match_any(text, _KW_FULFILL_NO):
        return '否'
    return '悬置'


def classify_interaction(text: str) -> str:
    if _match_any(text, _KW_INTERACT_QUERY):
        return '质疑'
    if _match_any(text, _KW_INTERACT_MEME):
        return '玩梗'
    if _match_any(text, _KW_INTERACT_SUPPLEMENT):
        return '补证'
    return '附和'


def like_range(n: int) -> str:
    if n >= 101:
        return '101+'
    if n >= 11:
        return '11-100'
    if n >= 1:
        return '1-10'
    return '0'


def reply_range(n: int) -> str:
    if n >= 21:
        return '21+'
    if n >= 6:
        return '6-20'
    if n >= 1:
        return '1-5'
    return '0'


def classify_position(c: dict) -> str:
    if c.get('stick_position', 0) or c.get('is_hot', 0):
        return '热评'
    if (c.get('digg_count') or 0) >= 1000:
        return '热评'
    return '普通'


def anonymize_text(text: str) -> str:
    """基础脱敏：去掉 @提及 与账号引用（人名/地名需人工复核）。"""
    t = text or ''
    t = re.sub(r'@[\w\u4e00-\u9fff-]+', '', t)   # @xxx 提及
    t = re.sub(r'回复\s*@[\w\u4e00-\u9fff-]+\s*的评论', '', t)
    return t.strip()


# ── 读取搜索记录 ──────────────────────────────────────────────────────────
def read_search_rows():
    wb = openpyxl.load_workbook(EXCEL_IN, data_only=True)
    ws = wb['每日搜索记录']
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None or not r[0]:
            continue
        sid, collect_date, entry, keyword, notes, link = r[0], r[1], r[2], r[3], r[4], r[5]
        raw_link = (link or '').strip()
        # 从整段分享文案中提取干净的抖音链接
        m = re.search(r'https://v\.douyin\.com/[A-Za-z0-9_\-]+/?', raw_link)
        if not m:
            m = re.search(r'https://www\.douyin\.com/video/\d+', raw_link)
        if not m:
            continue
        link = m.group(0)
        rows.append({
            'search_day_id': str(sid),
            'collect_date': norm_date(collect_date),
            'search_entry': (entry or '').strip(),
            'search_keyword': (keyword or '').strip(),
            'link': link,
        })
    return rows


# ── 主流程 ────────────────────────────────────────────────────────────────
def main():
    auth = load_env()
    print(f'读取 Excel: {EXCEL_IN}')
    rows = read_search_rows()
    print(f'共 {len(rows)} 条搜索记录待采集\n')

    video_records = []   # 视频记录 sheet 行
    comment_records = []  # 评论摘录 sheet 行
    vid_idx = 0
    cid_idx = 0

    for i, row in enumerate(rows, 1):
        short = row['link']
        print(f"[{i}/{len(rows)}] {row['search_day_id']} | 关键词={row['search_keyword']} | {short}")

        try:
            work_url, aweme_id = resolve_short_url(short)
        except Exception as e:
            print(f'    短链接解析失败: {e}')
            continue
        print(f'    → aweme_id={aweme_id}')

        # 视频详情
        detail = fetch_work_info(auth, work_url)
        if not detail:
            print('    aweme_detail 为空，跳过')
            continue

        desc = (detail.get('desc') or '').strip()
        author = detail.get('author') or {}
        statistics = detail.get('statistics') or {}
        comment_count = statistics.get('comment_count', 0)
        publish_date = ts_to_date(detail.get('create_time'))

        account_type = classify_account(author)

        vid_idx += 1
        video_id = f'V-DY-{vid_idx:04d}'
        video_records.append([
            video_id,
            row['collect_date'],
            row['search_entry'],
            row['search_keyword'],
            work_url,
            row['search_keyword'] or (desc[:30] if desc else ''),
            account_type,
            publish_date,
            f'评论{comment_count}条' + ('，热评多' if comment_count >= 100 else ''),
            '',  # selection_reason 待人工
            '不记录用户名/头像/主页等可识别信息',
        ])
        print(f'    视频: {desc[:40]!r} | 账号={account_type} | 发布={publish_date} | 评论={comment_count}')

        # 评论抓取
        comments = get_comments_limited(auth, work_url, MAX_COMMENTS_PER_VIDEO)
        print(f'    抓取评论 {len(comments)} 条')

        for c in comments:
            raw = (c.get('text') or '').strip()
            if not raw:
                continue
            cid_idx += 1
            corpus_id = f'C-DY-{cid_idx:04d}'
            anon = anonymize_text(raw)
            dg = c.get('digg_count') or 0
            rp = c.get('reply_comment_total') or 0
            narrative = classify_narrative(raw)
            comment_records.append([
                corpus_id,
                video_id,
                row['collect_date'],
                raw,
                anon,
                classify_position(c),
                like_range(dg),
                reply_range(rp),
                narrative,
                classify_narrator(raw),
                ('真的/亲眼/老人说' if _match_any(raw, _KW_EVIDENCE) else ''),
                classify_omen(raw),
                classify_fulfillment(raw),
                ('有' if _match_any(raw, _KW_MORAL) else '无'),
                classify_interaction(raw),
                '需改写',  # use_in_paper 默认需人工复核匿名化
                '',
            ])

        time.sleep(REQUEST_INTERVAL)

    # 写回 Excel
    print(f'\n写回结果: {EXCEL_OUT}')
    print(f'  视频记录 {len(video_records)} 条，评论摘录 {len(comment_records)} 条')
    write_result(video_records, comment_records)
    print('完成。')


def write_result(video_records, comment_records):
    wb = openpyxl.load_workbook(EXCEL_IN)

    # 视频记录 sheet
    ws_v = wb['视频记录']
    header_row = 1
    # 删除示例行（第2行）及之后所有旧数据
    if ws_v.max_row > header_row:
        ws_v.delete_rows(header_row + 1, ws_v.max_row - header_row)
    for row in video_records:
        ws_v.append(row)

    # 评论摘录 sheet
    ws_c = wb['评论摘录']
    if ws_c.max_row > 1:
        ws_c.delete_rows(2, ws_c.max_row - 1)
    for row in comment_records:
        ws_c.append(row)

    wb.save(EXCEL_OUT)


if __name__ == '__main__':
    main()
